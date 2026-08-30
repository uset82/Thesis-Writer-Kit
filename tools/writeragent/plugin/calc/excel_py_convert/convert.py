# SPDX-License-Identifier: GPL-3.0-or-later
"""Orchestrate Excel ↔ DAG-style ``=PY`` conversion. Details in ``to_dag.py``."""

from __future__ import annotations

import json
import logging
import re
import zipfile
from pathlib import Path
from typing import TYPE_CHECKING, Any
from xml.etree import ElementTree as ET  # nosemgrep: use-defused-xml  # local .xlsx ZIP parts

from plugin.calc.excel_py_convert.parse_dag_formulas import iter_dag_py_formulas_xlsx
from plugin.calc.excel_py_convert.parse_excel_ooxml import (
    _findall,
    _find_child,
    _local,
    _workbook_sheets,
    load_excel_model,
)
from plugin.calc.excel_py_convert.script_bank import CODE_SHEET_PREFIX, iter_a1_span
from plugin.calc.excel_py_convert.to_dag import convert_model_to_dag
from plugin.calc.excel_py_convert.to_excel import (
    assign_script_bank,
    convert_dag_cells_to_excel,
    convert_dag_report_to_excel,
    deps_for_xlws_export,
    python_scripts_xml,
    xlws_py_formula,
)

if TYPE_CHECKING:
    from plugin.calc.excel_py_convert.models import ConversionReport, ConvertedCell

log = logging.getLogger(__name__)

# Udprop: per-cell JSON so auto-save can restore return_type / data_args / excel_deps.
EXCEL_PY_DAG_META_PROP = "ExcelPyDagMeta"
# Optional same payload inside a DAG .xlsx (CLI without --from-report). Kept implemented but
# off by default — product path uses in-memory udprop; flip to True if needed later.
PACKAGE_META_PART = "xl/writeragentExcelPyMeta.json"
USE_PACKAGE_META = False

_SSML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_RE_A1 = re.compile(r"^([A-Za-z]+)(\d+)$")
_CONTENT_TYPES_PY_OVERRIDE = (
    '<Override PartName="/xl/pythonScripts.xml" ContentType="application/xml"/>'
)


def convert_to_dag(path: str | Path, *, best_effort: bool = False) -> ConversionReport:
    """Excel XLSX or JSON fixture → DAG-style conversion report."""
    model = load_excel_model(path)
    return convert_model_to_dag(model, best_effort=best_effort)


def convert_to_excel(path: str | Path) -> ConversionReport:
    """Workbook with DAG ``=PY`` formulas → Excel report (``xl(%Pn%)`` + script indices).

    Pair with ``write_excel_python_xlsx`` to emit native ``pythonScripts.xml`` /
    ``_xlfn._xlws.PY``. Resolves ``py_code_*`` bank refs when reading ``.xlsx``.

    For fidelity (return_type / data_args), prefer ``convert_dag_report_to_excel`` on
    an in-memory DAG report, or pass cell meta via udprop ``ExcelPyDagMeta``.
    """
    path = Path(path)
    if path.suffix.lower() == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict) and data.get("direction") == "dag":
            from plugin.calc.excel_py_convert.models import ConversionReport as _CR

            return convert_dag_report_to_excel(_CR.from_dict(data))
        return convert_dag_cells_to_excel(_triples_from_json(data), report_meta=data if isinstance(data, dict) else None)
    items = iter_dag_py_formulas_xlsx(path)
    package_meta = load_package_meta(path) if USE_PACKAGE_META else {}
    if package_meta:
        rebuilt: list[tuple[str, str, str, dict[str, Any]]] = []
        for sheet, cell, formula in items:
            meta: dict[str, Any] = {}
            stored = package_meta.get(cell_meta_key(sheet, cell))
            if isinstance(stored, dict):
                meta = dict(stored)
            rebuilt.append((sheet, cell, formula, meta))
        report = convert_dag_cells_to_excel(rebuilt)
    else:
        report = convert_dag_cells_to_excel(items)
    report.source_path = str(path)
    return report


def cell_meta_key(sheet: str, cell: str) -> str:
    return f"{sheet}!{cell}"


def dag_report_to_meta_payload(report: ConversionReport) -> dict[str, Any]:
    """Compact per-cell meta for udprop / ``PACKAGE_META_PART`` (auto-save / CLI reverse).

    Includes ``excel_deps`` for Table/ANCHORARRAY export fidelity — see models module doc.
    """
    out: dict[str, Any] = {}
    for c in report.cells:
        if not c.converted:
            continue
        out[cell_meta_key(c.sheet, c.cell)] = {
            "return_type": int(c.return_type or 0),
            "data_args": list(c.data_args),
            "excel_deps": list(c.excel_deps),
            "ordering_args": list(c.ordering_args),
            "array_ref": c.array_ref,
            "bindings": [
                {
                    "a1": b.a1,
                    "header_mode": b.header_mode,
                    "role": b.role,
                    "original_indices": list(b.original_indices),
                }
                for b in c.bindings
            ],
        }
    return out


def write_package_meta(out_path: Path, report: ConversionReport) -> None:
    """Embed ``PACKAGE_META_PART`` JSON into a DAG xlsx (after formula rewrite).

    Not called when ``USE_PACKAGE_META`` is False (default). Kept for a possible
    later CLI/offline path; prefer ``--from-report`` / udprop for now.
    """
    payload = dag_report_to_meta_payload(report)
    blob = json.dumps(payload, separators=(",", ":")).encode("utf-8")
    tmp = out_path.with_suffix(out_path.suffix + ".tmpmeta")
    with zipfile.ZipFile(out_path, "r") as zin, zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            if info.filename == PACKAGE_META_PART:
                continue
            zout.writestr(info, zin.read(info.filename))
        zout.writestr(PACKAGE_META_PART, blob)
    tmp.replace(out_path)


def load_package_meta(path: str | Path) -> dict[str, Any]:
    """Load ``PACKAGE_META_PART`` from an xlsx, or ``{}``.

    Callers should respect ``USE_PACKAGE_META`` before relying on this.
    """
    path = Path(path)
    try:
        with zipfile.ZipFile(path, "r") as zin:
            if PACKAGE_META_PART not in zin.namelist():
                return {}
            data = json.loads(zin.read(PACKAGE_META_PART).decode("utf-8"))
    except (OSError, zipfile.BadZipFile, json.JSONDecodeError, UnicodeError):
        return {}
    return data if isinstance(data, dict) else {}


def store_dag_meta_on_doc(doc: Any, report: ConversionReport) -> None:
    """Persist DAG conversion meta on *doc* for later native Excel export on save."""
    from plugin.doc.udprops import set_document_property

    payload = dag_report_to_meta_payload(report)
    set_document_property(doc, EXCEL_PY_DAG_META_PROP, json.dumps(payload, separators=(",", ":")))


def load_dag_meta_from_doc(doc: Any) -> dict[str, Any]:
    """Load ``ExcelPyDagMeta`` JSON map, or ``{}``."""
    from plugin.doc.udprops import get_document_property

    raw = get_document_property(doc, EXCEL_PY_DAG_META_PROP)
    if not raw:
        return {}
    try:
        data = json.loads(str(raw))
    except (TypeError, json.JSONDecodeError):
        return {}
    return data if isinstance(data, dict) else {}


def _triples_from_json(data: Any) -> list[tuple[str, str, str, dict[str, Any]]]:
    if isinstance(data, dict) and "cells" in data:
        out: list[tuple[str, str, str, dict[str, Any]]] = []
        for c in data["cells"]:
            formula = c.get("dag_formula") or c.get("formula")
            if not formula:
                continue
            out.append((str(c.get("sheet", "Sheet1")), str(c.get("cell", "A1")), str(formula), dict(c)))
        return out
    if isinstance(data, list):
        return [(str(x["sheet"]), str(x["cell"]), str(x["formula"]), dict(x)) for x in data]
    raise ValueError("JSON must be a dag report or list of {sheet, cell, formula}")


def convert_path(
    path: str | Path,
    *,
    direction: str,
    out_report: str | Path | None = None,
    best_effort: bool = False,
    from_report: str | Path | None = None,
) -> ConversionReport:
    """Convert *path* in *direction* ``dag`` or ``excel``; optionally write JSON report.

    When ``direction == "excel"`` and *from_report* is a DAG conversion JSON,
    use ``convert_dag_report_to_excel`` (preserves return_type / data_args).
    """
    direction = direction.strip().lower()
    if direction == "dag":
        report = convert_to_dag(path, best_effort=best_effort)
    elif direction == "excel":
        if from_report is not None:
            from plugin.calc.excel_py_convert.models import ConversionReport as _CR

            data = json.loads(Path(from_report).read_text(encoding="utf-8"))
            report = convert_dag_report_to_excel(_CR.from_dict(data))
            report.source_path = report.source_path or str(path)
        else:
            report = convert_to_excel(path)
    else:
        raise ValueError("direction must be 'dag' or 'excel'")
    if out_report is not None:
        Path(out_report).write_text(json.dumps(report.to_dict(), indent=2) + "\n", encoding="utf-8")
    return report


def _xlsx_formula_for_cell(cell: ConvertedCell) -> str:
    """Render a comma-separated OOXML ``=PY`` formula (script bank ref, no Calc sanitizer)."""
    from plugin.calc.excel_py_convert.script_bank import formula_for_converted_cell

    return formula_for_converted_cell(cell, separator=",", excel_escape=True, use_script_bank=True)


def _clear_spill_range(ws: Any, anchor: str, array_ref: str) -> None:
    """Clear cached/array result cells in *array_ref*, keeping the anchor for rewrite."""
    if not array_ref:
        return
    cells = iter_a1_span(array_ref)
    if len(cells) <= 1:
        return
    for coord in cells:
        if coord == anchor:
            continue
        try:
            ws[coord].value = None
        except Exception:
            continue


def _strip_python_in_excel_parts(out_path: Path) -> None:
    """Remove obsolete Python-in-Excel package parts after formula rewrite."""
    drop_prefixes = (
        "xl/pythonScripts",
        "xl/python",
    )
    drop_exact = {
        "xl/pythonScripts.xml",
    }
    tmp = out_path.with_suffix(out_path.suffix + ".tmpstrip")
    with zipfile.ZipFile(out_path, "r") as zin, zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            name = info.filename
            if name in drop_exact or any(name.startswith(p) for p in drop_prefixes):
                continue
            # Drop content-type / rel entries are left; orphan rels are harmless enough.
            # Controlled: also strip workbook rel targeting pythonScripts.
            data = zin.read(name)
            if name.endswith(".rels") or name == "[Content_Types].xml":
                text = data.decode("utf-8", errors="ignore")
                if "pythonScripts" in text or ("python" in text.lower() and "Override" in text):
                    # Remove lines referencing pythonScripts
                    lines = []
                    for line in text.splitlines(keepends=True):
                        if "pythonScripts" in line or "pythonScript" in line:
                            continue
                        if 'PartName="/xl/python' in line:
                            continue
                        lines.append(line)
                    data = "".join(lines).encode("utf-8")
            zout.writestr(info, data)
    tmp.replace(out_path)


def write_dag_formulas_xlsx(
    source_xlsx: str | Path,
    report: ConversionReport,
    out_path: str | Path,
    *,
    strip_python_parts: bool = True,
) -> None:
    """Copy *source_xlsx* and replace successfully converted PY cells with DAG formulas.

    - Parks rewritten Python on visible ``py_code_<Sheet>`` sheets at the **same A1**
      as each caller when ``len(code) > 1000``; shorter scripts stay inline in ``=PY("…")``.
    - OOXML formulas use **comma** separators (not Calc ``;``).
    - Clears the source array/spill ``ref`` range (except the anchor) so old
      cached results do not block the new spill.
    - Fails closed on unmapped sheet titles (no silent first-sheet fallback).
    - Optionally strips ``xl/pythonScripts.xml`` and related package parts.
    """
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import IllegalCharacterError

    from plugin.calc.excel_py_convert.script_bank import (
        collect_script_bank,
        report_safety_warnings,
        write_script_bank_openpyxl,
    )

    source_xlsx = Path(source_xlsx)
    out_path = Path(out_path)
    wb = load_workbook(source_xlsx)
    sheet_by_key = {ws.title: ws for ws in wb.worksheets}
    errors: list[str] = []

    bank, bank_warnings = collect_script_bank(report)
    for w in bank_warnings:
        log.warning("excel_py convert: %s", w)
    for w in report_safety_warnings(report):
        log.warning("excel_py convert: %s", w)
    write_script_bank_openpyxl(wb, bank)

    for cell in report.cells:
        if not cell.converted or not cell.converted_code:
            continue
        ws = sheet_by_key.get(cell.sheet)
        if ws is None:
            # Accept sheet1 → first sheet only when the report used fixture aliases
            # AND there is exactly one worksheet — still prefer exact titles.
            lower = {t.lower(): w for t, w in sheet_by_key.items()}
            ws = lower.get(cell.sheet.lower())
        if ws is None:
            errors.append(f"unmapped sheet {cell.sheet!r} for cell {cell.cell}")
            continue
        if cell.array_ref:
            _clear_spill_range(ws, cell.cell, cell.array_ref)
        formula = _xlsx_formula_for_cell(cell)
        try:
            ws[cell.cell] = formula
        except IllegalCharacterError as exc:
            errors.append(f"{cell.sheet}!{cell.cell}: {exc}")

    if errors:
        wb.close()
        raise ValueError("write_dag_formulas_xlsx failed:\n" + "\n".join(errors))

    wb.save(out_path)
    wb.close()
    if strip_python_parts:
        _strip_python_in_excel_parts(out_path)
    if USE_PACKAGE_META:
        write_package_meta(out_path, report)


def _a1_row_col(a1: str) -> tuple[int, str] | None:
    m = _RE_A1.match((a1 or "").replace("$", "").strip())
    if not m:
        return None
    return int(m.group(2)), m.group(1).upper()


def _ensure_sheet_data(ws_root: ET.Element) -> ET.Element:
    for child in list(ws_root):
        if _local(child.tag) == "sheetData":
            return child
    tag = f"{{{_SSML_NS}}}sheetData" if ws_root.tag.startswith("{") else "sheetData"
    el = ET.SubElement(ws_root, tag)
    return el


def _ensure_cell(ws_root: ET.Element, a1: str) -> ET.Element:
    """Return ``<c r="A1">``, creating row/cell under sheetData when missing."""
    parsed = _a1_row_col(a1)
    if parsed is None:
        raise ValueError(f"invalid A1: {a1!r}")
    row_n, col_letters = parsed
    a1_u = f"{col_letters}{row_n}"
    sheet_data = _ensure_sheet_data(ws_root)
    row_el: ET.Element | None = None
    for row in list(sheet_data):
        if _local(row.tag) != "row":
            continue
        if (row.attrib.get("r") or "") == str(row_n):
            row_el = row
            break
        for c in list(row):
            if _local(c.tag) == "c" and (c.attrib.get("r") or "").replace("$", "") == a1_u:
                return c
    if row_el is None:
        row_tag = f"{{{_SSML_NS}}}row" if sheet_data.tag.startswith("{") else "row"
        row_el = ET.SubElement(sheet_data, row_tag)
        row_el.set("r", str(row_n))
    for c in list(row_el):
        if _local(c.tag) == "c" and (c.attrib.get("r") or "").replace("$", "") == a1_u:
            return c
    c_tag = f"{{{_SSML_NS}}}c" if row_el.tag.startswith("{") else "c"
    cell = ET.SubElement(row_el, c_tag)
    cell.set("r", a1_u)
    return cell


def _set_cell_xlws_formula(ws_root: ET.Element, a1: str, formula: str, array_ref: str = "") -> None:
    cell = _ensure_cell(ws_root, a1)
    # Drop cached value / type so Excel recalculates from the formula.
    for child in list(cell):
        if _local(child.tag) in ("v", "is"):
            cell.remove(child)
    if "t" in cell.attrib:
        del cell.attrib["t"]
    f = _find_child(cell, "f")
    if f is None:
        f_tag = f"{{{_SSML_NS}}}f" if cell.tag.startswith("{") else "f"
        f = ET.SubElement(cell, f_tag)
    body = formula[1:] if formula.startswith("=") else formula
    f.text = body
    if array_ref:
        f.set("t", "array")
        f.set("ref", array_ref.replace("$", ""))
    else:
        if "t" in f.attrib:
            del f.attrib["t"]
        if "ref" in f.attrib:
            del f.attrib["ref"]


def _clear_spill_xml(ws_root: ET.Element, anchor: str, array_ref: str) -> None:
    if not array_ref:
        return
    for coord in iter_a1_span(array_ref):
        if coord == anchor.replace("$", ""):
            continue
        for c in _findall(ws_root, "c"):
            if (c.attrib.get("r") or "").replace("$", "") != coord:
                continue
            for child in list(c):
                if _local(child.tag) in ("f", "v", "is"):
                    c.remove(child)
            if "t" in c.attrib:
                del c.attrib["t"]


def _patch_content_types(data: bytes) -> bytes:
    text = data.decode("utf-8", errors="ignore")
    if "pythonScripts.xml" in text:
        return data
    # Insert Override before closing </Types>
    if "</Types>" not in text:
        return data
    inject = "  " + _CONTENT_TYPES_PY_OVERRIDE + "\n"
    return text.replace("</Types>", inject + "</Types>").encode("utf-8")


def _drop_py_code_sheets_from_workbook(wb_xml: bytes, rels_xml: bytes, drop_rids: set[str]) -> tuple[bytes, bytes]:
    """Remove ``py_code_*`` sheet entries from workbook.xml + workbook rels."""
    if not drop_rids:
        return wb_xml, rels_xml
    wb = ET.fromstring(wb_xml)
    for sheets_el in list(wb):
        if _local(sheets_el.tag) != "sheets":
            continue
        for sh in list(sheets_el):
            if _local(sh.tag) != "sheet":
                continue
            rid = ""
            for k, v in sh.attrib.items():
                if k.endswith("}id") or k in ("r:id", "id"):
                    rid = v
                    break
            title = sh.attrib.get("name") or ""
            if rid in drop_rids or title.startswith(CODE_SHEET_PREFIX):
                sheets_el.remove(sh)
    rels = ET.fromstring(rels_xml)
    for rel in list(rels):
        if _local(rel.tag) != "Relationship":
            continue
        if (rel.attrib.get("Id") or "") in drop_rids:
            rels.remove(rel)
    return ET.tostring(wb, encoding="utf-8", xml_declaration=True), ET.tostring(
        rels, encoding="utf-8", xml_declaration=True
    )


def write_excel_python_xlsx(
    source_xlsx: str | Path,
    report: ConversionReport,
    out_path: str | Path,
) -> None:
    """Write native Excel Python-in-Excel package (stdlib ZipFile only).

    - Banks ``converted_code`` into ``xl/pythonScripts.xml`` (``xl(%Pn%)`` bodies)
    - Sets each cell to ``_xlfn._xlws.PY(scriptIndex, returnType, deps…)``
    - Patches ``[Content_Types].xml``; strips Calc ``py_code_*`` sheets from the package
    """
    source_xlsx = Path(source_xlsx)
    out_path = Path(out_path)
    if report.direction != "excel":
        raise ValueError("write_excel_python_xlsx requires report.direction == 'excel'")
    scripts, _cells = assign_script_bank(report.cells)
    if not scripts or not any(c.converted and c.script_index >= 0 for c in report.cells):
        raise ValueError("write_excel_python_xlsx: no converted Excel PY cells to write")

    # sheet title → list of ConvertedCell
    by_sheet: dict[str, list[ConvertedCell]] = {}
    for cell in report.cells:
        if not cell.converted or cell.script_index < 0:
            continue
        by_sheet.setdefault(cell.sheet, []).append(cell)

    tmp = out_path.with_suffix(out_path.suffix + ".tmpexcelpy")
    errors: list[str] = []

    with zipfile.ZipFile(source_xlsx, "r") as zin:
        sheets = _workbook_sheets(zin)
        title_to_part = {s.title: s.part_name for s in sheets if s.part_name}
        title_to_part_l = {s.title.lower(): s.part_name for s in sheets if s.part_name}
        # Map py_code sheet parts for drop
        drop_parts: set[str] = set()
        drop_rids: set[str] = set()
        try:
            wb_root = ET.fromstring(zin.read("xl/workbook.xml"))
            for sheets_el in list(wb_root):
                if _local(sheets_el.tag) != "sheets":
                    continue
                for sh in list(sheets_el):
                    if _local(sh.tag) != "sheet":
                        continue
                    title = sh.attrib.get("name") or ""
                    if not title.startswith(CODE_SHEET_PREFIX):
                        continue
                    rid = ""
                    for k, v in sh.attrib.items():
                        if k.endswith("}id") or k in ("r:id", "id"):
                            rid = v
                            break
                    part = title_to_part.get(title) or title_to_part_l.get(title.lower(), "")
                    if part:
                        drop_parts.add(part)
                    if rid:
                        drop_rids.add(rid)
        except Exception:
            log.debug("excel_py write: py_code sheet discovery failed", exc_info=True)

        # Pre-parse worksheet roots we will patch
        patched: dict[str, ET.Element] = {}
        for sheet_title, cells in by_sheet.items():
            part = title_to_part.get(sheet_title) or title_to_part_l.get(sheet_title.lower(), "")
            if not part:
                errors.append(f"unmapped sheet {sheet_title!r}")
                continue
            try:
                root = patched.get(part) or ET.fromstring(zin.read(part))
            except KeyError:
                errors.append(f"missing worksheet part {part!r} for {sheet_title!r}")
                continue
            for cell in cells:
                try:
                    if cell.array_ref:
                        _clear_spill_xml(root, cell.cell, cell.array_ref)
                    formula = xlws_py_formula(
                        cell.script_index, cell.return_type, deps_for_xlws_export(cell)
                    )
                    _set_cell_xlws_formula(root, cell.cell, formula, array_ref=cell.array_ref)
                except Exception as exc:
                    errors.append(f"{cell.sheet}!{cell.cell}: {exc}")
            patched[part] = root

        if errors:
            raise ValueError("write_excel_python_xlsx failed:\n" + "\n".join(errors))

        scripts_bytes = python_scripts_xml(scripts)
        wb_xml = zin.read("xl/workbook.xml")
        rels_xml = zin.read("xl/_rels/workbook.xml.rels")
        if drop_rids:
            wb_xml, rels_xml = _drop_py_code_sheets_from_workbook(wb_xml, rels_xml, drop_rids)

        with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            written_scripts = False
            for info in zin.infolist():
                name = info.filename
                if name in drop_parts or name == PACKAGE_META_PART:
                    continue
                if name.startswith("xl/worksheets/_rels/") and any(
                    p.split("/")[-1] in name for p in drop_parts
                ):
                    # Drop sheet rels for removed py_code sheets when path matches.
                    base = name.rsplit("/", 1)[-1].replace(".rels", "")
                    if any(p.endswith(base) for p in drop_parts):
                        continue
                if name == "xl/pythonScripts.xml":
                    zout.writestr(info, scripts_bytes)
                    written_scripts = True
                    continue
                if name == "[Content_Types].xml":
                    data = _patch_content_types(zin.read(name))
                    # Also drop Override lines for removed worksheet parts
                    if drop_parts:
                        text = data.decode("utf-8", errors="ignore")
                        lines = []
                        for line in text.splitlines(keepends=True):
                            if any(f'PartName="/{p}"' in line or f'PartName="/{p.lstrip("/")}"' in line for p in drop_parts):
                                continue
                            lines.append(line)
                        data = "".join(lines).encode("utf-8")
                    zout.writestr(info, data)
                    continue
                if name == "xl/workbook.xml":
                    zout.writestr(info, wb_xml)
                    continue
                if name == "xl/_rels/workbook.xml.rels":
                    zout.writestr(info, rels_xml)
                    continue
                if name in patched:
                    root = patched[name]
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                    zout.writestr(info, data)
                    continue
                zout.writestr(info, zin.read(name))
            if not written_scripts:
                zout.writestr("xl/pythonScripts.xml", scripts_bytes)

    tmp.replace(out_path)


def convert_uno_doc_to_excel(doc: Any) -> ConversionReport:
    """Snapshot DAG ``=PY`` cells from an open Calc doc (UNO) → Excel conversion report.

    Resolves ``py_code_*`` bank cell strings in memory so long scripts export correctly.
    When ``ExcelPyDagMeta`` udprop is present (set on auto-open), uses stored
    ``return_type`` / ``data_args`` (legacy ``ordering_args`` ignored on export).
    """
    from plugin.calc.address_utils import index_to_column
    from plugin.calc.excel_py_convert.parse_dag_formulas import BANK_REF_RE
    from plugin.calc.excel_py_convert.script_bank import normalize_bank_a1
    from plugin.calc.python.cell_discovery import canonicalize_py_formula_for_parse, is_py_formula_text
    from plugin.calc.python.formula_edit import (
        CALC_PYTHON_FN,
        escape_code_for_excel_formula,
        parse_python_formula,
    )

    meta_map = load_dag_meta_from_doc(doc)
    triples: list[tuple[str, str, str, dict[str, Any]]] = []
    sheets = doc.getSheets()
    sheet_by_name: dict[str, Any] = {}
    for i in range(sheets.getCount()):
        sh = sheets.getByIndex(i)
        sheet_by_name[str(sh.getName())] = sh

    def _bank_string(code: str) -> str | None:
        m = BANK_REF_RE.match((code or "").strip())
        if not m:
            return None
        bank_sheet = sheet_by_name.get(m.group(1))
        if bank_sheet is None:
            lower = {k.lower(): v for k, v in sheet_by_name.items()}
            bank_sheet = lower.get(m.group(1).lower())
        if bank_sheet is None:
            return None
        try:
            a1 = normalize_bank_a1(m.group(2))
            return str(bank_sheet.getCellRangeByName(a1).getString() or "")
        except Exception:
            return None

    # Higher ceiling than sidebar discovery — export must see every PY cell.
    cell_flag_formula = 16
    max_scan = 100_000
    for sheet_name, sheet in sheet_by_name.items():
        if sheet_name.startswith(CODE_SHEET_PREFIX):
            continue
        try:
            formula_cells = sheet.queryContentCells(cell_flag_formula)
        except Exception:
            continue
        if formula_cells is None:
            continue
        try:
            count = int(formula_cells.getCount())
        except Exception:
            continue
        scanned = 0
        for i in range(count):
            if scanned >= max_scan:
                break
            try:
                cell_range = formula_cells.getByIndex(i)
                addr = cell_range.getRangeAddress()
                formula_matrix = cell_range.getFormulas() if hasattr(cell_range, "getFormulas") else None
            except Exception:
                continue
            if formula_matrix is None:
                continue
            for r_idx, row_formulas in enumerate(formula_matrix):
                row = addr.StartRow + r_idx
                for c_idx, formula in enumerate(row_formulas):
                    scanned += 1
                    if not formula or not is_py_formula_text(str(formula)):
                        continue
                    col = addr.StartColumn + c_idx
                    a1 = f"{index_to_column(col)}{row + 1}"
                    canonical = canonicalize_py_formula_for_parse(str(formula))
                    parts = parse_python_formula(canonical)
                    if parts is None:
                        continue
                    resolved = _bank_string(parts.code)
                    if resolved is not None:
                        formula_out = (
                            f'={CALC_PYTHON_FN}("{escape_code_for_excel_formula(resolved)}"{parts.data_suffix}'
                        )
                    else:
                        formula_out = canonical
                    cell_meta: dict[str, Any] = {}
                    stored = meta_map.get(cell_meta_key(sheet_name, a1))
                    if isinstance(stored, dict):
                        cell_meta = dict(stored)
                    triples.append((sheet_name, a1, formula_out, cell_meta))

    report = convert_dag_cells_to_excel(triples)
    return report
