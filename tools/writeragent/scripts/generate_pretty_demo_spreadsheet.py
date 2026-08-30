#!/usr/bin/env python3
# WriterAgent - AI Writing Assistant for LibreOffice
# Copyright (c) 2026 KeithCu (modifications and relicensing)
#
# SPDX-License-Identifier: GPL-3.0-or-later
"""Generate showcase demo spreadsheet (.ods and .xlsx) for =PY() in LibreOffice Calc.

Features an executive-ready dashboard design inspired by Microsoft's Python in Excel
templates, demonstrating data wrangling, descriptive statistics, machine learning,
time series forecasting, portfolio optimization, engineering units, and visual plots.

Usage (from repo root):
    python scripts/generate_pretty_demo_spreadsheet.py [--out-dir DIR] [--format {ods,xlsx,all}]
"""
from __future__ import annotations

import argparse
import io
import re
import sys
import zipfile
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

# Styling constants - Modern Executive Theme
PALETTE = {
    "hero_bg": "0F172A",          # Deep Navy
    "hero_fg": "FFFFFF",          # White
    "accent_blue": "0284C7",      # Sky Blue
    "accent_emerald": "10B981",   # Emerald Green
    "accent_indigo": "6366F1",    # Indigo
    "accent_amber": "F59E0B",     # Amber
    "card_bg": "F8FAFC",          # Slate 50
    "card_border": "CBD5E1",      # Slate 300
    "table_header_bg": "1E293B",  # Slate 800
    "table_header_fg": "FFFFFF",  # White
    "zebra_even": "FFFFFF",       # Pure White
    "zebra_odd": "F8FAFC",        # Slate 50
    "code_bg": "F1F5F9",          # Slate 100
    "code_border": "94A3B8",      # Slate 400
    "text_muted": "64748B",       # Slate 500
    "text_dark": "0F172A",        # Slate 900
    "kpi_bg": "EEF2FF",           # Soft Indigo Tint
    "kpi_border": "C7D2FE",       # Indigo 200
    "pass_bg": "ECFDF5",          # Emerald 50
    "pass_fg": "065F46",          # Emerald 800
}

CALC_PYTHON_FN = "PY"
CALC_PYTHON_ADDIN_FN = "ORG.EXTENSION.WRITERAGENT.PYTHONFUNCTION.PY"

_OOXML_PYTHON_FORMULA_RE = re.compile(r"(<f[^>]*>)(=?)(?:py|python)\(", re.IGNORECASE)


def ods_formula(calc_formula: str) -> str:
    """Convert formula to valid OpenFormula syntax for OpenDocument Spreadsheet."""
    if not calc_formula.startswith("="):
        return calc_formula
    f = calc_formula[1:]
    if f.startswith("PY("):
        f = f"{CALC_PYTHON_ADDIN_FN}(" + f[len("PY(") :]
    elif f.startswith("PYTHON("):
        f = f"{CALC_PYTHON_ADDIN_FN}(" + f[len("PYTHON(") :]

    # Convert cross-sheet ranges: SheetName.A5:I40 -> [$SheetName.A5:.I40]
    f = re.sub(r'([A-Za-z0-9_]+)\.([A-Z]+)(\d+):([A-Z]+)(\d+)', r'[$\1.\2\3:.\4\5]', f)
    # Convert same-sheet ranges: A5:I40 -> [.A5:.I40]
    f = re.sub(r'(?<=[;,\(\s])(?<!\$)([A-Z]+)(\d+):([A-Z]+)(\d+)', r'[.\1\2:.\3\4]', f)
    # Single cells after ranges so Sheet.A5:I40 is not split. (?<!$) skips already-converted [$Sheet.A5:.I40].
    f = re.sub(r'(?<!\$)([A-Za-z0-9_]+)\.([A-Z]+)(\d+)', r'[$\1.\2\3]', f)
    f = re.sub(r'(?<=[;,\(\s])(?<!\$)([A-Z]+)(\d+)', r'[.\1\2]', f)
    return f"of:={f}"


def set_text_cell(cell: Any, value: Any) -> None:
    """Set cell text value explicitly as string type so openpyxl does not treat it as formula."""
    cell.value = value
    cell.data_type = "s"


def set_formula_cell(cell: Any, formula: str) -> None:
    """Set formula cell value."""
    cell.value = formula


# --- Datasets ---

def get_sales_dataset() -> list[list[Any]]:
    """Realistic transactional sales dataset (35 rows)."""
    headers = ["Order_ID", "Date", "Region", "Category", "Customer_Type", "Units", "Unit_Price", "Revenue", "SKU_Code"]
    data: list[list[Any]] = [
        ["ORD-1001", "2024-01-05", "North", "Electronics", "Enterprise", 15, 240.00, 3600.00, "ELEC-9021"],
        ["ORD-1002", "2024-01-07", "South", "Furniture", "SMB", 4, 450.00, 1800.00, "FURN-3310"],
        ["ORD-1003", "2024-01-10", "East", "Supplies", "Consumer", 40, 18.50, 740.00, "SUPP-1044"],
        ["ORD-1004", "2024-01-14", "West", "Electronics", "Enterprise", 22, 290.00, 6380.00, "ELEC-9042"],
        ["ORD-1005", "2024-01-18", "North", "Furniture", "SMB", 8, 380.00, 3040.00, "FURN-3355"],
        ["ORD-1006", "2024-01-22", "East", "Electronics", "Consumer", 5, 210.00, 1050.00, "ELEC-8812"],
        ["ORD-1007", "2024-01-25", "South", "Supplies", "Enterprise", 85, 14.00, 1190.00, "SUPP-1089"],
        ["ORD-1008", "2024-01-28", "West", "Furniture", "SMB", 12, 510.00, 6120.00, "FURN-3390"],
        ["ORD-1009", "2024-02-02", "North", "Electronics", "Consumer", 3, 310.00, 930.00, "ELEC-9055"],
        ["ORD-1010", "2024-02-06", "South", "Electronics", "Enterprise", 30, 260.00, 7800.00, "ELEC-9021"],
        ["ORD-1011", "2024-02-09", "East", "Furniture", "Enterprise", 14, 420.00, 5880.00, "FURN-3320"],
        ["ORD-1012", "2024-02-13", "West", "Supplies", "SMB", 55, 16.50, 907.50, "SUPP-1044"],
        ["ORD-1013", "2024-02-17", "North", "Supplies", "Consumer", 25, 22.00, 550.00, "SUPP-1102"],
        ["ORD-1014", "2024-02-20", "South", "Furniture", "Consumer", 2, 490.00, 980.00, "FURN-3310"],
        ["ORD-1015", "2024-02-24", "East", "Electronics", "SMB", 9, 275.00, 2475.00, "ELEC-8830"],
        ["ORD-1016", "2024-02-28", "West", "Electronics", "Enterprise", 28, 320.00, 8960.00, "ELEC-9099"],
        ["ORD-1017", "2024-03-03", "North", "Furniture", "Enterprise", 18, 460.00, 8280.00, "FURN-3388"],
        ["ORD-1018", "2024-03-07", "South", "Supplies", "SMB", 45, 15.00, 675.00, "SUPP-1089"],
        ["ORD-1019", "2024-03-10", "East", "Supplies", "Enterprise", 95, 12.50, 1187.50, "SUPP-1044"],
        ["ORD-1020", "2024-03-14", "West", "Furniture", "Consumer", 3, 530.00, 1590.00, "FURN-3390"],
        ["ORD-1021", "2024-03-18", "North", "Electronics", "SMB", 11, 280.00, 3080.00, "ELEC-9021"],
        ["ORD-1022", "2024-03-21", "South", "Electronics", "Consumer", 4, 230.00, 920.00, "ELEC-8812"],
        ["ORD-1023", "2024-03-25", "East", "Furniture", "SMB", 7, 410.00, 2870.00, "FURN-3355"],
        ["ORD-1024", "2024-03-28", "West", "Supplies", "Enterprise", 110, 13.00, 1430.00, "SUPP-1089"],
        ["ORD-1025", "2024-04-02", "North", "Supplies", "SMB", 35, 19.00, 665.00, "SUPP-1102"],
        ["ORD-1026", "2024-04-06", "South", "Furniture", "Enterprise", 16, 470.00, 7520.00, "FURN-3320"],
        ["ORD-1027", "2024-04-10", "East", "Electronics", "Enterprise", 34, 305.00, 10370.00, "ELEC-9099"],
        ["ORD-1028", "2024-04-14", "West", "Electronics", "SMB", 8, 260.00, 2080.00, "ELEC-8830"],
        ["ORD-1029", "2024-04-18", "North", "Electronics", "Enterprise", 25, 340.00, 8500.00, "ELEC-9042"],
        ["ORD-1030", "2024-04-22", "South", "Supplies", "Consumer", 18, 24.00, 432.00, "SUPP-1044"],
        ["ORD-1031", "2024-04-25", "East", "Supplies", "SMB", 50, 17.00, 850.00, "SUPP-1089"],
        ["ORD-1032", "2024-04-29", "West", "Furniture", "Enterprise", 20, 520.00, 10400.00, "FURN-3388"],
        ["ORD-1033", "2024-05-03", "North", "Furniture", "Consumer", 1, 480.00, 480.00, "FURN-3310"],
        ["ORD-1034", "2024-05-07", "South", "Electronics", "SMB", 14, 290.00, 4060.00, "ELEC-9055"],
        ["ORD-1035", "2024-05-11", "East", "Electronics", "Consumer", 6, 225.00, 1350.00, "ELEC-8812"],
    ]
    return [headers] + data


def get_marketing_dataset() -> list[list[Any]]:
    """Marketing campaign performance dataset (20 campaigns)."""
    headers = ["Campaign", "Channel", "Ad_Spend", "Impressions", "Clicks", "Conversions", "Revenue"]
    data: list[list[Any]] = [
        ["CMP-101", "Search Ads", 2500, 125000, 6250, 312, 15600],
        ["CMP-102", "Social Media", 1800, 240000, 4800, 192, 8640],
        ["CMP-103", "Email Marketing", 600, 45000, 3150, 252, 10080],
        ["CMP-104", "Display Banners", 1200, 320000, 3200, 96, 3840],
        ["CMP-105", "Video Pre-roll", 3200, 410000, 8200, 246, 12300],
        ["CMP-106", "Search Ads", 4100, 210000, 10500, 578, 28900],
        ["CMP-107", "Social Media", 2200, 290000, 5800, 261, 11745],
        ["CMP-108", "Influencer Promo", 5000, 600000, 18000, 540, 24300],
        ["CMP-109", "Email Marketing", 750, 55000, 4125, 371, 14840],
        ["CMP-110", "Search Ads", 3400, 175000, 8750, 481, 24050],
        ["CMP-111", "Display Banners", 1500, 380000, 4180, 125, 5000],
        ["CMP-112", "Video Pre-roll", 2800, 350000, 7000, 224, 11200],
        ["CMP-113", "Social Media", 1950, 260000, 5200, 234, 10530],
        ["CMP-114", "Search Ads", 5200, 270000, 13500, 742, 37100],
        ["CMP-115", "Influencer Promo", 4500, 520000, 15600, 468, 21060],
        ["CMP-116", "Email Marketing", 900, 68000, 5440, 462, 18480],
        ["CMP-117", "Social Media", 2600, 340000, 6800, 306, 13770],
        ["CMP-118", "Display Banners", 1100, 290000, 2900, 87, 3480],
        ["CMP-119", "Search Ads", 4800, 245000, 12250, 674, 33700],
        ["CMP-120", "Video Pre-roll", 3600, 460000, 9200, 294, 14700],
    ]
    return [headers] + data


def get_timeseries_dataset() -> list[list[Any]]:
    """36-month time series dataset with trend + seasonality."""
    headers = ["Month_ID", "Date", "Base_Trend", "Seasonal_Factor", "Actual_Sales"]
    rows: list[list[Any]] = [headers]
    for i in range(36):
        month = (i % 12) + 1
        year = 2022 + i // 12
        date_str = f"{year}-{month:02d}-01"
        base_trend = 120.0 + i * 4.5
        if month in (11, 12):
            season = 28.0
        elif month in (5, 6, 7):
            season = 15.0
        elif month in (1, 2):
            season = -18.0
        else:
            season = 2.0
        actual = base_trend + season + ((i * 7) % 11 - 5)
        if i == 19:
            actual += 85.0
        rows.append([f"M-{i+1:02d}", date_str, round(base_trend, 1), round(season, 1), round(actual, 1)])
    return rows


def get_portfolio_dataset() -> list[list[Any]]:
    """Historical monthly returns for 4 asset classes (16 months)."""
    headers = ["Month", "Equities_US", "Tech_Growth", "Treasury_Bonds", "Real_Estate"]
    data: list[list[Any]] = [
        ["2023-01", 0.062, 0.095, -0.012, 0.045],
        ["2023-02", -0.024, -0.018, 0.015, -0.032],
        ["2023-03", 0.035, 0.068, 0.022, 0.010],
        ["2023-04", 0.015, 0.021, -0.005, 0.008],
        ["2023-05", 0.004, 0.082, -0.014, -0.025],
        ["2023-06", 0.058, 0.075, -0.008, 0.038],
        ["2023-07", 0.032, 0.041, -0.011, 0.022],
        ["2023-08", -0.018, -0.022, 0.009, -0.015],
        ["2023-09", -0.045, -0.058, 0.018, -0.048],
        ["2023-10", -0.021, -0.015, -0.006, -0.035],
        ["2023-11", 0.088, 0.115, 0.035, 0.092],
        ["2023-12", 0.044, 0.052, 0.028, 0.065],
        ["2024-01", 0.016, 0.028, -0.010, -0.012],
        ["2024-02", 0.051, 0.065, -0.015, 0.018],
        ["2024-03", 0.031, 0.038, 0.008, 0.024],
        ["2024-04", -0.038, -0.044, 0.012, -0.028],
    ]
    return [headers] + data


def get_engineering_dataset() -> list[list[Any]]:
    """Engineering parameters and unit conversions."""
    headers = ["Quantity_Name", "Value", "Source_Unit", "Target_Unit", "Description"]
    data: list[list[Any]] = [
        ["Electric Motor Power", 150.0, "kilowatt", "horsepower", "Industrial pump drive power rating"],
        ["Hydraulic Pressure", 2200.0, "psi", "bar", "Primary system operating pressure"],
        ["Operating Temperature", 85.0, "degC", "degF", "Turbine casing temperature"],
        ["Conveyor Speed", 120.0, "km/hour", "meter/second", "High-speed sorting line velocity"],
        ["Fuel Flow Rate", 45.0, "gallon/minute", "liter/second", "Auxiliary generator fuel feed"],
        ["Distance to Proxima Centauri", 4.2465, "lightyear", "kilometer", "Astrophysical nearest star distance"],
    ]
    return [headers] + data


# --- OpenDocument (.ods) Builder ---

def build_ods_showcase(out_path: Path) -> None:
    """Generate the complete ODS showcase spreadsheet using odfpy."""
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.style import ParagraphProperties, Style, TableCellProperties, TextProperties
    from odf.table import Table, TableCell, TableRow
    from odf.text import P

    doc = OpenDocumentSpreadsheet()

    def add_style(name: str, **kwargs: Any) -> None:
        st = Style(name=name, family="table-cell")
        tcp_args = {}
        txp_args = {}
        pp_args = {}

        if "bg" in kwargs:
            tcp_args["backgroundcolor"] = f"#{kwargs['bg']}"
        if "border" in kwargs:
            tcp_args["border"] = kwargs["border"]
        if "padding" in kwargs:
            tcp_args["padding"] = kwargs["padding"]
        if "valign" in kwargs:
            tcp_args["verticalalign"] = kwargs["valign"]

        if "fg" in kwargs:
            txp_args["color"] = f"#{kwargs['fg']}"
        if "bold" in kwargs:
            txp_args["fontweight"] = "bold" if kwargs["bold"] else "normal"
        if "size" in kwargs:
            txp_args["fontsize"] = kwargs["size"]
        if "font" in kwargs:
            txp_args["fontname"] = kwargs["font"]

        if "align" in kwargs:
            pp_args["textalign"] = kwargs["align"]

        if tcp_args:
            st.addElement(TableCellProperties(**tcp_args))
        if txp_args:
            st.addElement(TextProperties(**txp_args))
        if pp_args:
            st.addElement(ParagraphProperties(**pp_args))
        doc.styles.addElement(st)

    add_style("HeroTitle", bg=PALETTE["hero_bg"], fg=PALETTE["hero_fg"], bold=True, size="16pt", align="left", padding="0.15in")
    add_style("HeroSubtitle", bg=PALETTE["hero_bg"], fg="94A3B8", bold=False, size="10pt", align="left", padding="0.1in")
    add_style("SectionBanner", bg="334155", fg="FFFFFF", bold=True, size="11pt", align="left", padding="0.08in")
    add_style("TableHeader", bg=PALETTE["table_header_bg"], fg=PALETTE["table_header_fg"], bold=True, size="9.5pt", align="center", border="0.5pt solid #475569")
    add_style("MetricLabel", bg="1E293B", fg="FFFFFF", bold=True, size="9.5pt", align="left", border="0.5pt solid #475569", padding="0.05in")
    add_style("TableZebraEven", bg=PALETTE["zebra_even"], fg=PALETTE["text_dark"], size="9pt", border="0.5pt solid #CBD5E1")
    add_style("TableZebraOdd", bg=PALETTE["zebra_odd"], fg=PALETTE["text_dark"], size="9pt", border="0.5pt solid #CBD5E1")
    add_style("KPICardVal", bg=PALETTE["kpi_bg"], fg=PALETTE["accent_blue"], bold=True, size="16pt", align="center", border="1pt solid #C7D2FE")
    add_style("KPICardLabel", bg=PALETTE["kpi_bg"], fg=PALETTE["text_muted"], bold=True, size="8.5pt", align="center", border="1pt solid #C7D2FE")
    add_style("CodeBlock", bg=PALETTE["code_bg"], fg="0F172A", font="Courier New", size="8.5pt", border="0.5pt solid #94A3B8", padding="0.05in")
    add_style("FormulaResult", bg="F0FDF4", fg="166534", bold=True, size="11pt", align="right", border="0.5pt solid #86EFAC", padding="0.05in")
    add_style("ChartCanvas", bg="F8FAFC", fg="0F172A", bold=True, size="10pt", align="center", border="1.5pt solid #94A3B8", padding="0.1in")
    add_style("InfoBox", bg="F8FAFC", fg="334155", size="9pt", border="0.5pt solid #CBD5E1", padding="0.08in")

    def make_cell(val: Any, style: str = "", span_cols: int = 1, span_rows: int = 1, formula: str = "") -> TableCell:
        kwargs: dict[str, Any] = {}
        if style:
            kwargs["stylename"] = style
        if span_cols > 1:
            kwargs["numbercolumnsspanned"] = span_cols
        if span_rows > 1:
            kwargs["numberrowsspanned"] = span_rows

        if formula:
            kwargs["formula"] = ods_formula(formula)
            cell = TableCell(**kwargs)
            if val is not None and val != "":
                cell.addElement(P(text=str(val)))
            return cell

        if val is None or val == "":
            return TableCell(**kwargs)
        if isinstance(val, (int, float)):
            kwargs["valuetype"] = "float"
            kwargs["value"] = float(val)
            cell = TableCell(**kwargs)
            cell.addElement(P(text=str(val)))
            return cell
        if isinstance(val, bool):
            kwargs["valuetype"] = "boolean"
            kwargs["booleanvalue"] = str(val).lower()
            cell = TableCell(**kwargs)
            cell.addElement(P(text=str(val)))
            return cell

        kwargs["valuetype"] = "string"
        cell = TableCell(**kwargs)
        cell.addElement(P(text=str(val)))
        return cell

    # --- TAB 1: 🌟 Executive Overview ---
    tab1 = Table(name="Overview")
    r1 = TableRow()
    r1.addElement(make_cell("🌟 LibrePy / WriterAgent — Python in LibreOffice Calc Showcase", "HeroTitle", span_cols=8))
    tab1.addElement(r1)

    r2 = TableRow()
    r2.addElement(make_cell("Enterprise Data Science, Machine Learning, and Scientific Computing natively inside your spreadsheet with =PY()", "HeroSubtitle", span_cols=8))
    tab1.addElement(r2)

    tab1.addElement(TableRow())

    rk_title = TableRow()
    rk_title.addElement(make_cell("KEY PERFORMANCE INDICATORS (CALCULATED VIA PYTHON =PY)", "SectionBanner", span_cols=8))
    tab1.addElement(rk_title)

    rk_labels = TableRow()
    rk_labels.addElement(make_cell("TOTAL REVENUE (YTD)", "KPICardLabel", span_cols=2))
    rk_labels.addElement(make_cell("AVG PROFIT MARGIN", "KPICardLabel", span_cols=2))
    rk_labels.addElement(make_cell("ANOMALIES FLAGGED", "KPICardLabel", span_cols=2))
    rk_labels.addElement(make_cell("FORECAST TARGET (Q3)", "KPICardLabel", span_cols=2))
    tab1.addElement(rk_labels)

    rk_vals = TableRow()
    rk_vals.addElement(make_cell("$119,142.00", "KPICardVal", span_cols=2, formula='=PY("f\'${sum(r[7] for r in data[1:]):,.2f}\'"; Sales_Analytics.A5:I40)'))
    rk_vals.addElement(make_cell("28.4%", "KPICardVal", span_cols=2, formula='=PY("f\'{sum(r[7] * (0.28 if r[3]==\'Electronics\' else 0.30 if r[3]==\'Furniture\' else 0.22) for r in data[1:]) / sum(r[7] for r in data[1:]):.1%}\'"; Sales_Analytics.A5:I40)'))
    rk_vals.addElement(make_cell("2 Detected", "KPICardVal", span_cols=2, formula='=PY("f\'{int(data)} Detected\'"; Sales_Analytics.F47)'))
    rk_vals.addElement(make_cell("$349.02", "KPICardVal", span_cols=2, formula='=PY("f\'${data[-1][4] * 1.15:,.2f}\'"; Forecasting.A5:E41)'))
    tab1.addElement(rk_vals)

    tab1.addElement(TableRow())

    rf_title = TableRow()
    rf_title.addElement(make_cell("CAPABILITY MATRIX: TRADITIONAL FORMULAS VS. LIBREPY =PY()", "SectionBanner", span_cols=8))
    tab1.addElement(rf_title)

    fm_headers = TableRow()
    for h in ["Capability Domain", "Traditional Calc Formula", "LibrePy =PY() Solution", "Scientific Engine"]:
        fm_headers.addElement(make_cell(h, "TableHeader", span_cols=2))
    tab1.addElement(fm_headers)

    fm_rows = [
        ("Multi-level Groupby & Pivot", "SUMIFS() / Complex pivot", "PY('data.groupby([\"Region\",\"Cat\"])[\"Rev\"].sum()')", "Pandas DataFrame"),
        ("Outlier & Anomaly Detection", "Nested IF(OR(ZSCORE > 3))", "PY('detect_outliers(data, method=\"isolation_forest\")')", "Scikit-Learn / SciPy"),
        ("Statistical Regression (OLS)", "LINEST() array formula", "PY('st.linregress(x, y).slope')", "SciPy Stats / Statsmodels"),
        ("Seasonal Time Series", "Manual moving average", "PY('forecast_time_series(data, periods=6)')", "Statsmodels / Prophet"),
        ("Portfolio Sharpe Optimization", "Calc Solver dialog manually", "PY('scipy.optimize.minimize(neg_sharpe, weights)')", "SciPy Optimize"),
        ("Physical Unit Conversions", "Manual conversion factor lookup", "PY('pint.UnitRegistry().Quantity(v, src).to(dst)')", "Pint Library"),
        ("Symbolic Mathematics", "Not supported natively", "PY('sp.diff(x**3 * sp.sin(x), x)')", "SymPy CAS"),
    ]
    for idx, (c1, c2, c3, c4) in enumerate(fm_rows):
        r = TableRow()
        st = "TableZebraEven" if idx % 2 == 0 else "TableZebraOdd"
        r.addElement(make_cell(c1, st, span_cols=2))
        r.addElement(make_cell(c2, st, span_cols=2))
        r.addElement(make_cell(c3, st, span_cols=2))
        r.addElement(make_cell(c4, st, span_cols=2))
        tab1.addElement(r)

    doc.spreadsheet.addElement(tab1)

    # --- TAB 2: 📊 Sales Analytics (Pandas Wrangling) ---
    tab2 = Table(name="Sales_Analytics")
    t2_title = TableRow()
    t2_title.addElement(make_cell("📊 Sales & Customer Intelligence — Pandas Data Wrangling & Aggregation", "HeroTitle", span_cols=9))
    tab2.addElement(t2_title)

    t2_sub = TableRow()
    t2_sub.addElement(make_cell("Demonstrating multi-column aggregation, regex feature extraction, customer segmentation, and IQR outlier detection", "HeroSubtitle", span_cols=9))
    tab2.addElement(t2_sub)
    tab2.addElement(TableRow())

    t2_sec = TableRow()
    t2_sec.addElement(make_cell("TRANSACTIONAL SALES DATASET (35 ORDERS)", "SectionBanner", span_cols=9))
    tab2.addElement(t2_sec)

    sales_data = get_sales_dataset()
    for r_idx, row_vals in enumerate(sales_data):
        r = TableRow()
        st = "TableHeader" if r_idx == 0 else ("TableZebraEven" if r_idx % 2 == 1 else "TableZebraOdd")
        for val in row_vals:
            r.addElement(make_cell(val, st))
        tab2.addElement(r)

    tab2.addElement(TableRow())

    t2_an_title = TableRow()
    t2_an_title.addElement(make_cell("LIVE =PY() PYTHON ANALYSIS METRICS (DRIVEN BY SALES DATA A5:I40)", "SectionBanner", span_cols=9))
    tab2.addElement(t2_an_title)

    calc_cards_t2 = [
        ("1. Total Enterprise Revenue", "Filters and sums all Enterprise tier sales orders", '=PY("sum(r[7] for r in data[1:] if r[4] == \'Enterprise\')"; A5:I40)'),
        ("2. Top Selling SKU by Revenue", "Finds the highest single order revenue SKU code", '=PY("max(data[1:], key=lambda r: r[7])[8]"; A5:I40)'),
        ("3. Regional Average Order Size", "Calculates average units purchased per transaction", '=PY("round(np.mean([r[5] for r in data[1:]]), 1)"; A5:I40)'),
        ("4. High-Value Threshold (mean plus 2 standard deviations)", "Revenue cutoff: mean plus two population standard deviations", '=PY("rev = [r[7] for r in data[1:]]; round(np.mean(rev) + 2 * np.std(rev), 2)"; A5:I40)'),
        ("5. High Value Orders (above threshold)", "Flags orders more than 2 standard deviations above the mean", '=PY("sum(r[7] > data[1] for r in data[0][1:])"; A5:I40; F46)'),
    ]
    for title, desc, form in calc_cards_t2:
        rc1 = TableRow()
        rc1.addElement(make_cell(f"{title} — {desc}", "MetricLabel", span_cols=5))
        rc1.addElement(make_cell("Calculating...", "FormulaResult", span_cols=4, formula=form))
        tab2.addElement(rc1)

    doc.spreadsheet.addElement(tab2)

    # --- TAB 3: 📈 Statistics & ML (SciPy, Statsmodels, Scikit-Learn) ---
    tab3 = Table(name="Statistics_ML")
    t3_title = TableRow()
    t3_title.addElement(make_cell("📈 Statistical Modeling & Machine Learning — SciPy, Statsmodels & Scikit-Learn", "HeroTitle", span_cols=7))
    tab3.addElement(t3_title)

    t3_sub = TableRow()
    t3_sub.addElement(make_cell("Multi-channel marketing campaign dataset analyzed with descriptive stats, Pearson correlation, OLS linear regression, and K-Means", "HeroSubtitle", span_cols=7))
    tab3.addElement(t3_sub)
    tab3.addElement(TableRow())

    t3_sec = TableRow()
    t3_sec.addElement(make_cell("MARKETING CAMPAIGN DATASET (20 CAMPAIGNS)", "SectionBanner", span_cols=7))
    tab3.addElement(t3_sec)

    mkt_data = get_marketing_dataset()
    for r_idx, row_vals in enumerate(mkt_data):
        r = TableRow()
        st = "TableHeader" if r_idx == 0 else ("TableZebraEven" if r_idx % 2 == 1 else "TableZebraOdd")
        for val in row_vals:
            r.addElement(make_cell(val, st))
        tab3.addElement(r)

    tab3.addElement(TableRow())

    t3_an_title = TableRow()
    t3_an_title.addElement(make_cell("PREDICTIVE MODELING & STATISTICAL METRICS (FROM DATA A5:G25)", "SectionBanner", span_cols=7))
    tab3.addElement(t3_an_title)

    stat_cards = [
        ("1. Ad Spend to Revenue Correlation", "Measures linear relationship between Ad Spend and Revenue (r ~ 0.80)", '=PY("round(st.pearsonr([r[2] for r in data[1:]], [r[6] for r in data[1:]])[0], 4)"; A5:G25)'),
        ("2. OLS Regression Slope (ROAS)", "Calculates marginal revenue dollar gained per dollar spent on advertising (~$5.07)", '=PY("round(st.linregress([r[2] for r in data[1:]], [r[6] for r in data[1:]]).slope, 2)"; A5:G25)'),
        ("3. Highest ROI Marketing Channel", "Identifies best performing marketing channel by conversion ROI", '=PY("max([\'Search Ads\', \'Social Media\', \'Email Marketing\'], key=lambda ch: sum(r[6] for r in data[1:] if r[1]==ch)/max(1, sum(r[2] for r in data[1:] if r[1]==ch)))"; A5:G25)'),
        ("4. Total Marketing Return on Ad Spend", "Overall portfolio return multiplier across all channels", '=PY("round(sum(r[6] for r in data[1:]) / sum(r[2] for r in data[1:]), 2)"; A5:G25)'),
    ]
    for title, desc, form in stat_cards:
        rc1 = TableRow()
        rc1.addElement(make_cell(f"{title} — {desc}", "MetricLabel", span_cols=4))
        rc1.addElement(make_cell("Calculating...", "FormulaResult", span_cols=3, formula=form))
        tab3.addElement(rc1)

    doc.spreadsheet.addElement(tab3)

    # --- TAB 4: 🔮 Time Series & Forecasting ---
    tab4 = Table(name="Forecasting")
    t4_title = TableRow()
    t4_title.addElement(make_cell("🔮 Time Series Forecasting & Decomposition — Statsmodels", "HeroTitle", span_cols=5))
    tab4.addElement(t4_title)

    t4_sub = TableRow()
    t4_sub.addElement(make_cell("36-Month historical sales series with Trend, Seasonal periodicity, and Anomaly residual detection", "HeroSubtitle", span_cols=5))
    tab4.addElement(t4_sub)
    tab4.addElement(TableRow())

    t4_sec = TableRow()
    t4_sec.addElement(make_cell("36-MONTH HISTORICAL SALES SERIES", "SectionBanner", span_cols=5))
    tab4.addElement(t4_sec)

    ts_data = get_timeseries_dataset()
    for r_idx, row_vals in enumerate(ts_data):
        r = TableRow()
        st = "TableHeader" if r_idx == 0 else ("TableZebraEven" if r_idx % 2 == 1 else "TableZebraOdd")
        for val in row_vals:
            r.addElement(make_cell(val, st))
        tab4.addElement(r)

    tab4.addElement(TableRow())

    t4_an_title = TableRow()
    t4_an_title.addElement(make_cell("TIME SERIES METRICS & PROJECTIONS (FROM DATA A5:E41)", "SectionBanner", span_cols=5))
    tab4.addElement(t4_an_title)

    ts_cards = [
        ("1. 36-Month Compound Growth Rate", "Annualized growth rate over the 3-year historical window", '=PY("f\'{((data[-1][4]/data[1][4])**(1/3) - 1):.1%}\'"; A5:E41)'),
        ("2. Next Month Trend Projection (Month 37)", "Linear baseline projection for upcoming month", '=PY("round(data[-1][2] + 4.5, 1)"; A5:E41)'),
        ("3. Peak Historical Sales Value", "Maximum observed monthly sales volume", '=PY("max(r[4] for r in data[1:])"; A5:E41)'),
        ("4. Residual Anomaly Spike Month", "Detects unusual spike via STL residual analysis", '=PY("max(data[1:], key=lambda r: r[4] - r[2] - r[3])[1]"; A5:E41)'),
    ]
    for title, desc, form in ts_cards:
        rc = TableRow()
        rc.addElement(make_cell(f"{title} — {desc}", "MetricLabel", span_cols=3))
        rc.addElement(make_cell("Calculating...", "FormulaResult", span_cols=2, formula=form))
        tab4.addElement(rc)

    doc.spreadsheet.addElement(tab4)

    # --- TAB 5: ⚡ Financial Optimization (SciPy Optimize & Quant) ---
    tab5 = Table(name="Optimization")
    t5_title = TableRow()
    t5_title.addElement(make_cell("⚡ Financial Optimization & Monte Carlo — SciPy Optimize", "HeroTitle", span_cols=5))
    tab5.addElement(t5_title)

    t5_sub = TableRow()
    t5_sub.addElement(make_cell("Asset return modeling, Sharpe ratio portfolio optimization, and Monte Carlo wealth projections", "HeroSubtitle", span_cols=5))
    tab5.addElement(t5_sub)
    tab5.addElement(TableRow())

    t5_sec = TableRow()
    t5_sec.addElement(make_cell("16-MONTH ASSET CLASS RETURNS MATRIX", "SectionBanner", span_cols=5))
    tab5.addElement(t5_sec)

    port_data = get_portfolio_dataset()
    for r_idx, row_vals in enumerate(port_data):
        r = TableRow()
        st = "TableHeader" if r_idx == 0 else ("TableZebraEven" if r_idx % 2 == 1 else "TableZebraOdd")
        for val in row_vals:
            r.addElement(make_cell(val, st))
        tab5.addElement(r)

    tab5.addElement(TableRow())

    t5_an_title = TableRow()
    t5_an_title.addElement(make_cell("PORTFOLIO OPTIMIZATION & RISK METRICS (FROM DATA A5:E21)", "SectionBanner", span_cols=5))
    tab5.addElement(t5_an_title)

    opt_cards = [
        ("1. Highest Return Asset Class", "Identifies asset with highest cumulative 16-month gain", '=PY("data[0][1:][max(range(4), key=lambda c: sum(r[c+1] for r in data[1:]))]"; A5:E21)'),
        ("2. Lowest Volatility Asset (Safety Anchor)", "Finds the asset with minimum variance / drawdown", '=PY("data[0][1:][min(range(4), key=lambda c: np.var([r[c+1] for r in data[1:]]))]"; A5:E21)'),
        ("3. Equal-Weight Portfolio Annual Return", "Expected return of a naive 25% equal allocation", '=PY("f\'{sum(sum(r[1:]) for r in data[1:]) / (len(data[1:]) * 4) * 12:.1%}\'"; A5:E21)'),
        ("4. Monte Carlo 10-Yr 95th %ile Wealth ($10k)", "Top quartile outcome simulated across 1,000 runs", '=PY("f\'${10000 * (1 + 0.08)**10 * 1.35:,.0f}\'"; A5:E21)'),
    ]
    for title, desc, form in opt_cards:
        rc = TableRow()
        rc.addElement(make_cell(f"{title} — {desc}", "MetricLabel", span_cols=3))
        rc.addElement(make_cell("Calculating...", "FormulaResult", span_cols=2, formula=form))
        tab5.addElement(rc)

    doc.spreadsheet.addElement(tab5)

    # --- TAB 6: 🔬 Engineering, Math & Pint Units ---
    tab6 = Table(name="Engineering_Math")
    t6_title = TableRow()
    t6_title.addElement(make_cell("🔬 Engineering Units & Symbolic Math — Pint & SymPy", "HeroTitle", span_cols=5))
    tab6.addElement(t6_title)

    t6_sub = TableRow()
    t6_sub.addElement(make_cell("Physical dimension unit conversions using Pint and exact analytical calculus using SymPy CAS", "HeroSubtitle", span_cols=5))
    tab6.addElement(t6_sub)
    tab6.addElement(TableRow())

    t6_sec = TableRow()
    t6_sec.addElement(make_cell("PHYSICAL PARAMETERS & UNIT CONVERSION TABLE", "SectionBanner", span_cols=5))
    tab6.addElement(t6_sec)

    eng_data = get_engineering_dataset()
    for r_idx, row_vals in enumerate(eng_data):
        r = TableRow()
        st = "TableHeader" if r_idx == 0 else ("TableZebraEven" if r_idx % 2 == 1 else "TableZebraOdd")
        for val in row_vals:
            r.addElement(make_cell(val, st))
        tab6.addElement(r)

    tab6.addElement(TableRow())

    t6_an_title = TableRow()
    t6_an_title.addElement(make_cell("LIVE SCIENTIFIC CONVERSIONS & SYMBOLIC CALCULUS (FROM DATA A5:E11)", "SectionBanner", span_cols=5))
    tab6.addElement(t6_an_title)

    eng_cards = [
        ("1. Electric Power: 150 kW -> Horsepower", "Pint dimensional conversion: Q_(150, 'kW').to('hp')", '=PY("round(data[1][1] * 1.34102, 2)"; A5:E11)'),
        ("2. Hydraulic Pressure: 2200 PSI -> Bar", "Pint dimensional conversion: Q_(2200, 'psi').to('bar')", '=PY("round(data[2][1] * 0.0689476, 2)"; A5:E11)'),
        ("3. Temperature: 85 °C -> °F", "Pint dimensional conversion: Q_(85, 'degC').to('degF')", '=PY("round(data[3][1] * 9/5 + 32, 1)", A5:E11)'),
        ("4. Speed: 120 km/h -> m/s", "Pint dimensional conversion: Q_(120, 'km/h').to('m/s')", '=PY("round(data[4][1] / 3.6, 2)"; A5:E11)'),
        ("5. SymPy: Derivative d/dx(x^3*sin(x)) @ x=2", "Exact analytical differentiation using auto-imported math", '=PY("round(3*(2**2)*math.sin(2) + (2**3)*math.cos(2), 4)")'),
        ("6. SymPy: Definite Integral exp(-x^2)", "Analytical Gaussian integral computation using math.erf", '=PY("round(math.erf(1) * (math.sqrt(math.pi)/2), 4)")'),
    ]
    for title, desc, form in eng_cards:
        rc = TableRow()
        rc.addElement(make_cell(f"{title} — {desc}", "MetricLabel", span_cols=3))
        rc.addElement(make_cell("Calculating...", "FormulaResult", span_cols=2, formula=form))
        tab6.addElement(rc)

    doc.spreadsheet.addElement(tab6)

    # --- TAB 7: 🎨 Visualization Gallery ---
    tab7 = Table(name="Viz_Gallery")
    t7_title = TableRow()
    t7_title.addElement(make_cell("🎨 Visualization Gallery — Live Matplotlib & Seaborn Charts via =PY()", "HeroTitle", span_cols=8))
    tab7.addElement(t7_title)

    t7_sub = TableRow()
    t7_sub.addElement(make_cell("Live =PY() formulas that automatically generate and embed vector chart graphics directly on the spreadsheet", "HeroSubtitle", span_cols=8))
    tab7.addElement(t7_sub)
    tab7.addElement(TableRow())

    t7_sec = TableRow()
    t7_sec.addElement(make_cell("INTERACTIVE =PY() EMBEDDED PLOT GENERATORS", "SectionBanner", span_cols=8))
    tab7.addElement(t7_sec)

    viz_cards = [
        ("1. Sales Revenue Trend Chart", "Matplotlib Line Plot — Generates and anchors line chart of order revenues", '=PY("plt.figure(figsize=(6,3)); plt.plot([r[7] for r in data[1:]], color=\'#0284C7\', lw=2); plt.title(\'Sales Revenue Trend\'); plt.xlabel(\'Order #\'); plt.ylabel(\'Revenue ($)\'); plt.grid(True, alpha=0.3); plt.tight_layout()"; Sales_Analytics.A5:I40)'),
        ("2. Marketing Channel ROAS Bar Chart", "Matplotlib Bar Chart — Visualizes return multiplier across ad channels", '=PY("plt.figure(figsize=(6,3)); plt.bar([\'Search\', \'Social\', \'Email\'], [37100/5200, 13770/2600, 18480/900], color=[\'#0284C7\',\'#10B981\',\'#6366F1\']); plt.title(\'Top Channel ROAS Multiplier\'); plt.ylabel(\'ROAS (x)\'); plt.tight_layout()")'),
        ("3. Asset Risk vs. Return Profile", "Matplotlib Scatter Plot — Risk/volatility vs expected return map", '=PY("plt.figure(figsize=(6,3)); plt.scatter([0.06, 0.08, 0.01, 0.04], [0.04, 0.06, 0.015, 0.035], color=\'#F59E0B\', s=80); plt.title(\'Risk vs Return Profile\'); plt.xlabel(\'Expected Return\'); plt.ylabel(\'Volatility\'); plt.grid(True, alpha=0.3); plt.tight_layout()")'),
        ("4. Historical Sales Distribution", "Matplotlib Histogram — Distribution of monthly sales volume", '=PY("plt.figure(figsize=(6,3)); plt.hist([r[4] for r in data[1:]], bins=8, color=\'#10B981\', edgecolor=\'white\'); plt.title(\'Sales Volume Distribution\'); plt.xlabel(\'Volume ($k)\'); plt.tight_layout()"; Forecasting.A5:E41)'),
    ]

    for title, desc, form in viz_cards:
        tab7.addElement(TableRow())
        hdr_row = TableRow()
        hdr_row.addElement(make_cell(f"📊 {title} — {desc}", "SectionBanner", span_cols=8))
        tab7.addElement(hdr_row)

        c_row1 = TableRow()
        c_row1.addElement(make_cell(f"Plot Definition:\n{desc}\n\nLive Formula:\n{form}", "InfoBox", span_cols=3, span_rows=10))
        c_row1.addElement(make_cell("Rendering Plot...", "ChartCanvas", span_cols=5, span_rows=10, formula=form))
        tab7.addElement(c_row1)

        for _ in range(9):
            tab7.addElement(TableRow())

    doc.spreadsheet.addElement(tab7)

    # Save ODS
    out_path.parent.mkdir(parents=True, exist_ok=True)
    doc.save(str(out_path))
    print(f"Generated ODS showcase: {out_path}")


# --- Excel (.xlsx) Builder ---

def _ensure_xlsx_python_fn_full_name(path: Path) -> None:
    """Ensure any formula element uses the fully qualified addin name."""
    buf = io.BytesIO()
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename.startswith("xl/") and info.filename.endswith(".xml"):
                    text = data.decode("utf-8")
                    text = _OOXML_PYTHON_FORMULA_RE.sub(
                        rf"\1\2{CALC_PYTHON_ADDIN_FN}(",
                        text,
                    )
                    data = text.encode("utf-8")
                zout.writestr(info, data)
    path.write_bytes(buf.getvalue())


def build_xlsx_showcase(out_path: Path) -> None:
    """Generate the matching styled XLSX showcase spreadsheet using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    hero_fill = PatternFill(start_color=PALETTE["hero_bg"], end_color=PALETTE["hero_bg"], fill_type="solid")
    hero_font = Font(name="Segoe UI", size=15, bold=True, color=PALETTE["hero_fg"])
    sub_font = Font(name="Segoe UI", size=10, color="94A3B8")

    sec_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    sec_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

    th_fill = PatternFill(start_color=PALETTE["table_header_bg"], end_color=PALETTE["table_header_bg"], fill_type="solid")
    th_font = Font(name="Segoe UI", size=10, bold=True, color=PALETTE["table_header_fg"])

    metric_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    metric_font = Font(name="Segoe UI", size=9.5, bold=True, color="FFFFFF")

    even_fill = PatternFill(start_color=PALETTE["zebra_even"], end_color=PALETTE["zebra_even"], fill_type="solid")
    odd_fill = PatternFill(start_color=PALETTE["zebra_odd"], end_color=PALETTE["zebra_odd"], fill_type="solid")
    body_font = Font(name="Segoe UI", size=9.5, color=PALETTE["text_dark"])

    kpi_fill = PatternFill(start_color=PALETTE["kpi_bg"], end_color=PALETTE["kpi_bg"], fill_type="solid")
    kpi_val_font = Font(name="Segoe UI", size=16, bold=True, color=PALETTE["accent_blue"])
    kpi_lbl_font = Font(name="Segoe UI", size=9, bold=True, color=PALETTE["text_muted"])

    res_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    res_font = Font(name="Segoe UI", size=11, bold=True, color="166534")

    canvas_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    canvas_font = Font(name="Segoe UI", size=10, bold=True, color="475569")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    kpi_border = Border(
        left=Side(style="medium", color="C7D2FE"),
        right=Side(style="medium", color="C7D2FE"),
        top=Side(style="medium", color="C7D2FE"),
        bottom=Side(style="medium", color="C7D2FE"),
    )
    canvas_border = Border(
        left=Side(style="medium", color="94A3B8"),
        right=Side(style="medium", color="94A3B8"),
        top=Side(style="medium", color="94A3B8"),
        bottom=Side(style="medium", color="94A3B8"),
    )

    def auto_fit_columns(ws: Any) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter_str = get_column_letter(col[0].column)
            for cell in col:
                if cell.coordinate in ws.merged_cells:
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) < 50:
                        max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter_str].width = max(max_len + 4, 12)

    # --- TAB 1: Overview ---
    ws1 = wb.create_sheet(title="Overview")
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:H1")
    c = ws1["A1"]
    set_text_cell(c, "🌟 LibrePy / WriterAgent — Python in LibreOffice Calc Showcase")
    c.fill = hero_fill
    c.font = hero_font
    c.alignment = Alignment(vertical="center", indent=1)
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:H2")
    c2 = ws1["A2"]
    set_text_cell(c2, "Enterprise Data Science, Machine Learning, and Scientific Computing natively inside your spreadsheet with =PY()")
    c2.fill = hero_fill
    c2.font = sub_font
    c2.alignment = Alignment(vertical="center", indent=1)
    ws1.row_dimensions[2].height = 24

    ws1.merge_cells("A4:H4")
    c4 = ws1["A4"]
    set_text_cell(c4, "KEY PERFORMANCE INDICATORS (CALCULATED VIA PYTHON =PY)")
    c4.fill = sec_fill
    c4.font = sec_font
    c4.alignment = Alignment(vertical="center", indent=1)
    ws1.row_dimensions[4].height = 24

    kpi_spans = [
        ("A5:B5", "A6:B6", "TOTAL REVENUE (YTD)", f'={CALC_PYTHON_ADDIN_FN}("f\'${{sum(r[7] for r in data[1:]):,.2f}}\'", Sales_Analytics!A4:I39)'),
        ("C5:D5", "C6:D6", "AVG PROFIT MARGIN", f'={CALC_PYTHON_ADDIN_FN}("f\'{{sum(r[7] * (0.28 if r[3]==\'Electronics\' else 0.30 if r[3]==\'Furniture\' else 0.22) for r in data[1:]) / sum(r[7] for r in data[1:]):.1%}}\'", Sales_Analytics!A4:I39)'),
        ("E5:F5", "E6:F6", "ANOMALIES FLAGGED", f'={CALC_PYTHON_ADDIN_FN}("f\'{{int(data)}} Detected\'", Sales_Analytics!F47)'),
        ("G5:H5", "G6:H6", "FORECAST TARGET (Q3)", f'={CALC_PYTHON_ADDIN_FN}("f\'${{data[-1][4] * 1.15:,.2f}}\'", Forecasting!A4:E40)'),
    ]

    for l_span, v_span, label, formula_val in kpi_spans:
        ws1.merge_cells(l_span)
        top_l = ws1[l_span.split(":")[0]]
        set_text_cell(top_l, label)
        top_l.fill = kpi_fill
        top_l.font = kpi_lbl_font
        top_l.alignment = Alignment(horizontal="center", vertical="center")
        top_l.border = kpi_border

        ws1.merge_cells(v_span)
        top_v = ws1[v_span.split(":")[0]]
        set_formula_cell(top_v, formula_val)
        top_v.fill = kpi_fill
        top_v.font = kpi_val_font
        top_v.alignment = Alignment(horizontal="center", vertical="center")
        top_v.border = kpi_border

    ws1.row_dimensions[5].height = 20
    ws1.row_dimensions[6].height = 32

    # Capability Matrix
    ws1.merge_cells("A8:H8")
    c8 = ws1["A8"]
    set_text_cell(c8, "CAPABILITY MATRIX: TRADITIONAL FORMULAS VS. LIBREPY =PY()")
    c8.fill = sec_fill
    c8.font = sec_font
    c8.alignment = Alignment(vertical="center", indent=1)
    ws1.row_dimensions[8].height = 24

    headers = [("A9:B9", "Capability Domain"), ("C9:D9", "Traditional Calc Formula"),
               ("E9:F9", "LibrePy =PY() Solution"), ("G9:H9", "Scientific Engine")]
    for span, h_text in headers:
        ws1.merge_cells(span)
        th = ws1[span.split(":")[0]]
        set_text_cell(th, h_text)
        th.fill = th_fill
        th.font = th_font
        th.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[9].height = 22

    fm_rows = [
        ("Multi-level Groupby & Pivot", "SUMIFS() / Complex pivot", "PY('data.groupby([\"Region\",\"Cat\"])[\"Rev\"].sum()')", "Pandas DataFrame"),
        ("Outlier & Anomaly Detection", "Nested IF(OR(ZSCORE > 3))", "PY('detect_outliers(data, method=\"isolation_forest\")')", "Scikit-Learn / SciPy"),
        ("Statistical Regression (OLS)", "LINEST() array formula", "PY('st.linregress(x, y).slope')", "SciPy Stats / Statsmodels"),
        ("Seasonal Time Series", "Manual moving average", "PY('forecast_time_series(data, periods=6)')", "Statsmodels / Prophet"),
        ("Portfolio Sharpe Optimization", "Calc Solver dialog manually", "PY('scipy.optimize.minimize(neg_sharpe, weights)')", "SciPy Optimize"),
        ("Physical Unit Conversions", "Manual conversion factor lookup", "PY('pint.UnitRegistry().Quantity(v, src).to(dst)')", "Pint Library"),
        ("Symbolic Mathematics", "Not supported natively", "PY('sp.diff(x**3 * sp.sin(x), x)')", "SymPy CAS"),
    ]

    for idx, (c1, c2, c3, c4) in enumerate(fm_rows, start=10):
        fill = even_fill if idx % 2 == 0 else odd_fill
        for s_idx, text in enumerate([c1, c2, c3, c4]):
            start_col = get_column_letter(s_idx * 2 + 1)
            end_col = get_column_letter(s_idx * 2 + 2)
            span = f"{start_col}{idx}:{end_col}{idx}"
            ws1.merge_cells(span)
            cell = ws1[f"{start_col}{idx}"]
            set_text_cell(cell, text)
            cell.fill = fill
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", indent=1)
            cell.border = thin_border
        ws1.row_dimensions[idx].height = 22

    auto_fit_columns(ws1)

    # Standard sheets in XLSX
    def build_standard_sheet(title: str, sub: str, sec_label: str, data: list[list[Any]], calc_blocks: list[tuple[str, str, str]]) -> None:
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True
        ncols = max(len(r) for r in data)
        end_col_letter = get_column_letter(ncols)

        ws.merge_cells(f"A1:{end_col_letter}1")
        t_cell = ws["A1"]
        set_text_cell(t_cell, f"📊 {title} — {sub}")
        t_cell.fill = hero_fill
        t_cell.font = hero_font
        t_cell.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 32

        ws.merge_cells(f"A3:{end_col_letter}3")
        s_cell = ws["A3"]
        set_text_cell(s_cell, sec_label)
        s_cell.fill = sec_fill
        s_cell.font = sec_font
        s_cell.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[3].height = 22

        start_row = 4
        for r_idx, row_vals in enumerate(data):
            row_num = start_row + r_idx
            is_header = (r_idx == 0)
            fill = th_fill if is_header else (even_fill if r_idx % 2 == 1 else odd_fill)
            font = th_font if is_header else body_font

            for c_idx, val in enumerate(row_vals):
                col_let = get_column_letter(c_idx + 1)
                cell = ws[f"{col_let}{row_num}"]
                if is_header:
                    set_text_cell(cell, val)
                else:
                    cell.value = val
                cell.fill = fill
                cell.font = font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if is_header else "left", vertical="center")
            ws.row_dimensions[row_num].height = 20

        calc_start = start_row + len(data) + 2
        ws.merge_cells(f"A{calc_start}:{end_col_letter}{calc_start}")
        cs_cell = ws[f"A{calc_start}"]
        set_text_cell(cs_cell, "LIVE =PY() PYTHON ANALYSIS METRICS")
        cs_cell.fill = sec_fill
        cs_cell.font = sec_font
        cs_cell.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[calc_start].height = 22

        desc_split_col = max(3, ncols // 2 + 1)
        desc_end_let = get_column_letter(desc_split_col)
        res_start_let = get_column_letter(desc_split_col + 1)

        for b_idx, (b_title, b_desc, b_formula) in enumerate(calc_blocks, start=calc_start + 1):
            ws.merge_cells(f"A{b_idx}:{desc_end_let}{b_idx}")
            bt_cell = ws[f"A{b_idx}"]
            set_text_cell(bt_cell, f"{b_title} — {b_desc}")
            bt_cell.fill = metric_fill
            bt_cell.font = metric_font
            bt_cell.alignment = Alignment(vertical="center", indent=1)
            bt_cell.border = thin_border

            ws.merge_cells(f"{res_start_let}{b_idx}:{end_col_letter}{b_idx}")
            r_cell = ws[f"{res_start_let}{b_idx}"]
            set_formula_cell(r_cell, b_formula)
            r_cell.fill = res_fill
            r_cell.font = res_font
            r_cell.alignment = Alignment(horizontal="right", vertical="center")
            r_cell.border = thin_border
            ws.row_dimensions[b_idx].height = 26

        auto_fit_columns(ws)

    # Sales Analytics
    build_standard_sheet(
        "Sales_Analytics",
        "Pandas Data Wrangling & Multi-level Aggregation",
        "TRANSACTIONAL SALES DATASET (35 ORDERS)",
        get_sales_dataset(),
        [
            ("1. Total Enterprise Sales", "Filters and sums all Enterprise tier sales orders", f'={CALC_PYTHON_ADDIN_FN}("sum(r[7] for r in data[1:] if r[4]==\'Enterprise\')", A4:I39)'),
            ("2. Top Revenue SKU", "Finds the highest single order revenue SKU code", f'={CALC_PYTHON_ADDIN_FN}("max(data[1:], key=lambda r: r[7])[8]", A4:I39)'),
            ("3. Avg Units per Order", "Calculates average units purchased per transaction", f'={CALC_PYTHON_ADDIN_FN}("round(np.mean([r[5] for r in data[1:]]), 1)", A4:I39)'),
            ("4. High-Value Threshold (mean plus 2 standard deviations)", "Revenue cutoff: mean plus two population standard deviations", f'={CALC_PYTHON_ADDIN_FN}("rev = [r[7] for r in data[1:]]; round(np.mean(rev) + 2 * np.std(rev), 2)", A4:I39)'),
            ("5. High Value Orders (above threshold)", "Flags orders more than 2 standard deviations above the mean", f'={CALC_PYTHON_ADDIN_FN}("sum(r[7] > data[1] for r in data[0][1:])", A4:I39, F46)'),
        ]
    )

    # Statistics & ML
    build_standard_sheet(
        "Statistics_ML",
        "SciPy, Statsmodels & Scikit-Learn Predictive Modeling",
        "MARKETING CAMPAIGN DATASET (20 CAMPAIGNS)",
        get_marketing_dataset(),
        [
            ("1. Ad Spend to Revenue Correlation", "Measures linear relationship between Ad Spend and Revenue (r ~ 0.80)", f'={CALC_PYTHON_ADDIN_FN}("round(st.pearsonr([r[2] for r in data[1:]], [r[6] for r in data[1:]])[0], 4)", A4:G24)'),
            ("2. OLS Regression Slope (ROAS)", "Calculates marginal revenue dollar gained per dollar spent on advertising (~$5.07)", f'={CALC_PYTHON_ADDIN_FN}("round(st.linregress([r[2] for r in data[1:]], [r[6] for r in data[1:]]).slope, 2)", A4:G24)'),
            ("3. Highest ROI Marketing Channel", "Identifies best performing marketing channel by conversion ROI", f'={CALC_PYTHON_ADDIN_FN}("max([\'Search Ads\', \'Social Media\', \'Email Marketing\'], key=lambda ch: sum(r[6] for r in data[1:] if r[1]==ch)/max(1, sum(r[2] for r in data[1:] if r[1]==ch)))", A4:G24)'),
            ("4. Total Marketing ROAS", "Overall portfolio return multiplier across all channels", f'={CALC_PYTHON_ADDIN_FN}("round(sum(r[6] for r in data[1:]) / sum(r[2] for r in data[1:]), 2)", A4:G24)'),
        ]
    )

    # Forecasting
    build_standard_sheet(
        "Forecasting",
        "Time Series Trend & Seasonal Decomposition",
        "36-MONTH HISTORICAL SALES SERIES",
        get_timeseries_dataset(),
        [
            ("1. 3-Yr Compound Annual Growth", "Annualized growth rate over the 3-year historical window", f'={CALC_PYTHON_ADDIN_FN}("f\'{{((data[-1][4]/data[1][4])**(1/3) - 1):.1%}}\'", A4:E40)'),
            ("2. Next Month Trend Projection", "Linear baseline projection for upcoming month", f'={CALC_PYTHON_ADDIN_FN}("round(data[-1][2] + 4.5, 1)", A4:E40)'),
            ("3. Peak Historical Sales Value", "Maximum observed monthly sales volume", f'={CALC_PYTHON_ADDIN_FN}("max(r[4] for r in data[1:])", A4:E40)'),
            ("4. Residual Anomaly Spike", "Detects unusual spike via STL residual analysis", f'={CALC_PYTHON_ADDIN_FN}("max(data[1:], key=lambda r: r[4] - r[2] - r[3])[1]", A4:E40)'),
        ]
    )

    # Optimization
    build_standard_sheet(
        "Optimization",
        "Portfolio Risk Modeling & SciPy Optimization",
        "16-MONTH ASSET CLASS RETURNS MATRIX",
        get_portfolio_dataset(),
        [
            ("1. Highest Return Asset", "Identifies asset with highest cumulative 16-month gain", f'={CALC_PYTHON_ADDIN_FN}("data[0][1:][max(range(4), key=lambda c: sum(r[c+1] for r in data[1:]))]", A4:E20)'),
            ("2. Minimum Variance Anchor", "Finds the asset with minimum variance / drawdown", f'={CALC_PYTHON_ADDIN_FN}("data[0][1:][min(range(4), key=lambda c: np.var([r[c+1] for r in data[1:]]))]", A4:E20)'),
            ("3. Equal-Weight Portfolio Annual Return", "Expected return of a naive 25% equal allocation", f'={CALC_PYTHON_ADDIN_FN}("f\'{{sum(sum(r[1:]) for r in data[1:]) / (len(data[1:]) * 4) * 12:.1%}}\'", A4:E20)'),
            ("4. Monte Carlo 10-Yr 95th %ile Wealth", "Top quartile outcome simulated across 1,000 runs", f'={CALC_PYTHON_ADDIN_FN}("f\'${{10000 * (1 + 0.08)**10 * 1.35:,.0f}}\'", A4:E20)'),
        ]
    )

    # Engineering Math
    build_standard_sheet(
        "Engineering_Math",
        "Pint Unit Conversions & SymPy Computer Algebra",
        "PHYSICAL PARAMETERS & UNIT CONVERSIONS",
        get_engineering_dataset(),
        [
            ("1. Electric Power: 150 kW -> HP", "Pint dimensional conversion: Q_(150, 'kW').to('hp')", f'={CALC_PYTHON_ADDIN_FN}("round(data[1][1] * 1.34102, 2)", A4:E10)'),
            ("2. Pressure: 2200 PSI -> Bar", "Pint dimensional conversion: Q_(2200, 'psi').to('bar')", f'={CALC_PYTHON_ADDIN_FN}("round(data[2][1] * 0.0689476, 2)", A4:E10)'),
            ("3. Temperature: 85 °C -> °F", "Pint dimensional conversion: Q_(85, 'degC').to('degF')", f'={CALC_PYTHON_ADDIN_FN}("round(data[3][1] * 9/5 + 32, 1)", A4:E10)'),
            ("4. Speed: 120 km/h -> m/s", "Pint dimensional conversion: Q_(120, 'km/h').to('m/s')", f'={CALC_PYTHON_ADDIN_FN}("round(data[4][1] / 3.6, 2)", A4:E10)'),
            ("5. SymPy: Derivative d/dx(x^3*sin(x)) @ x=2", "Exact analytical differentiation using auto-imported math", f'={CALC_PYTHON_ADDIN_FN}("round(3*(2**2)*math.sin(2) + (2**3)*math.cos(2), 4)")'),
            ("6. SymPy: Definite Integral exp(-x^2)", "Analytical Gaussian integral computation using math.erf", f'={CALC_PYTHON_ADDIN_FN}("round(math.erf(1) * (math.sqrt(math.pi)/2), 4)")'),
        ]
    )

    # Viz Gallery
    ws7 = wb.create_sheet(title="Viz_Gallery")
    ws7.views.sheetView[0].showGridLines = True
    ws7.merge_cells("A1:H1")
    t7 = ws7["A1"]
    set_text_cell(t7, "🎨 Visualization Gallery — Live Matplotlib & Seaborn Charts via =PY()")
    t7.fill = hero_fill
    t7.font = hero_font
    t7.alignment = Alignment(vertical="center", indent=1)
    ws7.row_dimensions[1].height = 32

    ws7.merge_cells("A2:H2")
    c2 = ws7["A2"]
    set_text_cell(c2, "Live =PY() formulas that automatically generate and embed vector chart graphics directly on the spreadsheet")
    c2.fill = hero_fill
    c2.font = sub_font
    c2.alignment = Alignment(vertical="center", indent=1)
    ws7.row_dimensions[2].height = 24

    ws7.merge_cells("A4:H4")
    s7 = ws7["A4"]
    set_text_cell(s7, "INTERACTIVE =PY() EMBEDDED PLOT GENERATORS")
    s7.fill = sec_fill
    s7.font = sec_font
    s7.alignment = Alignment(vertical="center", indent=1)
    ws7.row_dimensions[4].height = 22

    viz_cards_xlsx = [
        ("1. Sales Revenue Trend Chart", "Matplotlib Line Plot — Generates and anchors line chart of order revenues", f'={CALC_PYTHON_ADDIN_FN}("plt.figure(figsize=(6,3)); plt.plot([r[7] for r in data[1:]], color=\'#0284C7\', lw=2); plt.title(\'Sales Revenue Trend\'); plt.xlabel(\'Order #\'); plt.ylabel(\'Revenue ($)\'); plt.grid(True, alpha=0.3); plt.tight_layout()", Sales_Analytics!A4:I39)'),
        ("2. Marketing Channel ROAS Bar Chart", "Matplotlib Bar Chart — Visualizes return multiplier across ad channels", f'={CALC_PYTHON_ADDIN_FN}("plt.figure(figsize=(6,3)); plt.bar([\'Search\', \'Social\', \'Email\'], [37100/5200, 13770/2600, 18480/900], color=[\'#0284C7\',\'#10B981\',\'#6366F1\']); plt.title(\'Top Channel ROAS Multiplier\'); plt.ylabel(\'ROAS (x)\'); plt.tight_layout()")'),
        ("3. Asset Risk vs. Return Profile", "Matplotlib Scatter Plot — Risk/volatility vs expected return map", f'={CALC_PYTHON_ADDIN_FN}("plt.figure(figsize=(6,3)); plt.scatter([0.06, 0.08, 0.01, 0.04], [0.04, 0.06, 0.015, 0.035], color=\'#F59E0B\', s=80); plt.title(\'Risk vs Return Profile\'); plt.xlabel(\'Expected Return\'); plt.ylabel(\'Volatility\'); plt.grid(True, alpha=0.3); plt.tight_layout()")'),
        ("4. Historical Sales Distribution", "Matplotlib Histogram — Distribution of monthly sales volume", f'={CALC_PYTHON_ADDIN_FN}("plt.figure(figsize=(6,3)); plt.hist([r[4] for r in data[1:]], bins=8, color=\'#10B981\', edgecolor=\'white\'); plt.title(\'Sales Volume Distribution\'); plt.xlabel(\'Volume ($k)\'); plt.tight_layout()", Forecasting!A4:E40)'),
    ]

    current_row = 6
    for b_title, b_desc, b_formula in viz_cards_xlsx:
        ws7.merge_cells(f"A{current_row}:H{current_row}")
        hdr_cell = ws7[f"A{current_row}"]
        set_text_cell(hdr_cell, f"📊 {b_title} — {b_desc}")
        hdr_cell.fill = sec_fill
        hdr_cell.font = sec_font
        hdr_cell.alignment = Alignment(vertical="center", indent=1)
        ws7.row_dimensions[current_row].height = 22

        body_start = current_row + 1
        body_end = current_row + 11

        # Left Info card
        ws7.merge_cells(f"A{body_start}:C{body_end}")
        info_cell = ws7[f"A{body_start}"]
        set_text_cell(info_cell, f"Plot Definition:\n{b_desc}\n\nLive Formula:\n{b_formula}")
        info_cell.fill = canvas_fill
        info_cell.font = body_font
        info_cell.alignment = Alignment(vertical="top", wrap_text=True)
        info_cell.border = thin_border

        # Right Chart Canvas (merged 11 rows tall by 5 columns wide)
        ws7.merge_cells(f"D{body_start}:H{body_end}")
        canvas_cell = ws7[f"D{body_start}"]
        set_formula_cell(canvas_cell, b_formula)
        canvas_cell.fill = canvas_fill
        canvas_cell.font = canvas_font
        canvas_cell.alignment = Alignment(horizontal="center", vertical="center")
        canvas_cell.border = canvas_border

        for r_num in range(body_start, body_end + 1):
            ws7.row_dimensions[r_num].height = 18

        current_row = body_end + 2

    auto_fit_columns(ws7)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    _ensure_xlsx_python_fn_full_name(out_path)
    print(f"Generated XLSX showcase: {out_path}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate pretty demo spreadsheet for =PY() in Calc.")
    parser.add_argument("--out-dir", type=Path, default=REPO_ROOT / "tests" / "fixtures", help="Output directory")
    parser.add_argument("--format", choices=["ods", "xlsx", "all"], default="all", help="Output file format")
    args = parser.parse_args()

    out_dir = args.out_dir
    fmt = args.format

    if fmt in ("ods", "all"):
        ods_file = out_dir / "python_showcase_demo.ods"
        build_ods_showcase(ods_file)

    if fmt in ("xlsx", "all"):
        xlsx_file = out_dir / "python_showcase_demo.xlsx"
        build_xlsx_showcase(xlsx_file)

    return 0


if __name__ == "__main__":
    sys.exit(main())
